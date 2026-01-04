from openpyxl import load_workbook
from openpyxl.styles import Side, Border

wb = load_workbook(filename='otput.xlsx')
sheet_name = wb.sheetnames[0]

ws1 = wb[sheet_name]
ws1.merge_cells('A5:B7')
ws1['A5'] = "あいうえお"

s = Side(style='thin')
b = Border(left=s, right=s, top=s, bottom=s)

cell = ws1['B2']
cell.border = b
wb.save('otput.xlsx')